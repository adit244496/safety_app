"""Seed EASE Score data from Excel files."""
import os
import openpyxl
from sqlalchemy.orm import Session
import models

CATEGORY_ROW = {
    "General Safety": 8,
    "Housekeeping": 26,
    "Personal Protective Equipment and work apparels": 42,
    "Work at Height": 49,
    "Electrical": 66,
    "Welding,  cutting and grinding": 84,
    "Vehicles, Earth movers & Lifting Equipment": 98,
    "Fire Prevention/ Protection": 109,
    "First Aid & Medical Facilities": 117,
    "Welfare Facilities and labour accomodation": 126,
    "Environment Aspects": 141,
    "Piling Work": 146,
    "Excavation and confined space job": 152,
    "Hand tools (manual and power driven)": 161,
}

MONTH_TO_NUM = {
    "JANUARY": 1, "FEBRUARY": 2, "MARCH": 3, "APRIL": 4,
    "MAY": 5, "JUNE": 6, "JULY": 7, "AUGUST": 8,
    "SEPTEMBER": 9, "OCTOBER": 10, "NOVEMBER": 11, "DECEMBER": 12,
}


def gradation_from_score(score: float) -> str:
    if score > 0.90:
        return "EXCELLENT"
    if score > 0.74:
        return "GOOD"
    if score >= 0.60:
        return "AVERAGE"
    return "BELOW AVERAGE"


def _parse_date_from_filename(filename: str):
    """Try to parse date range from filename like 'Date  16-02-26 to 27-02-26'."""
    import re
    m = re.search(r"(\d{2})-(\d{2})-(\d{2})\s+to\s+(\d{2})-(\d{2})-(\d{2})", filename)
    if m:
        d1, mo1, y1, d2, mo2, y2 = m.groups()
        return f"20{y1}-{mo1}-{d1}", f"20{y2}-{mo2}-{d2}"
    return None, None


def seed_ease_from_excel(db: Session, excel_path: str, sheet_name: str = "UTPALAA"):
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb[sheet_name]

        project_name = str(ws["C3"].value or "UNKNOWN").strip()
        month_str = str(ws["E3"].value or "JANUARY").upper().strip()
        year_val = ws["H3"].value
        period_year = int(year_val) if year_val else 2026
        period_month = MONTH_TO_NUM.get(month_str, 1)

        date_from, date_to = _parse_date_from_filename(os.path.basename(excel_path))

        overall_val = ws["C171"].value
        overall_score = float(overall_val) if isinstance(overall_val, (int, float)) else None
        ease_cat_val = ws["H171"].value
        ease_category = str(ease_cat_val).strip() if ease_cat_val else None

        for category, row_num in CATEGORY_ROW.items():
            h_val = ws.cell(row=row_num, column=8).value
            if h_val == "NA" or h_val is None:
                score = None
                gradation = "NA"
            elif isinstance(h_val, (int, float)):
                score = float(h_val)
                gradation = gradation_from_score(score)
            else:
                score = None
                gradation = "NA"

            entry = models.EaseScoreEntry(
                project_name=project_name,
                period_month=period_month,
                period_year=period_year,
                date_from=date_from,
                date_to=date_to,
                category=category,
                score=score,
                gradation=gradation,
                overall_score=overall_score,
                ease_category=ease_category,
            )
            db.add(entry)

        db.commit()
        print(f"EASE Score seeded: {project_name} {month_str} {period_year}")
    except Exception as exc:
        db.rollback()
        print(f"Warning: Could not seed EASE score from {excel_path}: {exc}")


def seed_ease_scores(db: Session):
    if db.query(models.EaseScoreEntry).first():
        return  # Already seeded

    base_dir = os.path.dirname(os.path.dirname(__file__))
    for fname in os.listdir(base_dir):
        lower = fname.lower()
        if lower.startswith("ease score") and lower.endswith(".xlsx"):
            fpath = os.path.join(base_dir, fname)
            seed_ease_from_excel(db, fpath, sheet_name="UTPALAA")
