from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from pydantic import BaseModel
from typing import Optional, List
import json
import io
import openpyxl
import threading
import logging
from difflib import get_close_matches
from sqlalchemy.orm import Session
from sqlalchemy import func, extract, case, nullslast
from datetime import datetime, date
from database import get_db, SessionLocal
import models
from auth import get_current_user, require_admin
from email_service import send_observation_email, build_observation_email

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/observations", tags=["observations"])


def calc_risk(severity: int, probability: int):
    factor = severity * probability
    level = "Low" if factor <= 4 else "Medium" if factor <= 12 else "High"
    return factor, level


def generate_obs_id(db: Session, project_id: int) -> str:
    project = db.query(models.Project).filter(models.Project.id == project_id).first()
    prefix = "".join(c for c in (project.name if project else "OBS") if c.isalnum()).upper()[:4]
    count = db.query(models.Observation).filter(models.Observation.project_id == project_id).count()
    date_str = datetime.now().strftime("%Y%m%d")
    seq = count + 1
    # Ensure uniqueness — walk forward if the ID is already taken (handles race conditions & deletions)
    while db.query(models.Observation).filter(
        models.Observation.observation_id == f"{prefix}-{date_str}-{str(seq).zfill(4)}"
    ).first() is not None:
        seq += 1
    return f"{prefix}-{date_str}-{str(seq).zfill(4)}"


def _assert_obs_access(obs: models.Observation, user: models.User, write_roles: Optional[List[str]] = None):
    """Raise 403 if user has no access to this observation's project, or lacks required role."""
    allowed = get_allowed_project_ids(user)
    if allowed is not None and obs.project_id not in allowed:
        raise HTTPException(403, "Access denied to this observation")
    if write_roles and user.role not in write_roles:
        raise HTTPException(403, "Your role cannot perform this action")


def get_allowed_project_ids(user: models.User) -> Optional[List[int]]:
    if user.role in ("SuperAdmin", "Admin", "PIC", "EIC"):
        return None
    # HO, PSO, Observer, Contractor — scoped to assigned projects
    project_ids = [up.project_id for up in user.user_projects]
    if not project_ids:
        return None  # no assignments = full access (admin intent: "all projects")
    return project_ids


def obs_to_dict(obs: models.Observation, db: Session) -> dict:
    images = db.query(models.ObservationImage).filter(models.ObservationImage.observation_id == obs.id).all()
    comments = db.query(models.ObservationComment).filter(models.ObservationComment.observation_id == obs.id).all()
    return {
        "id": obs.id,
        "observation_id": obs.observation_id,
        "project_id": obs.project_id,
        "project_name": obs.project.name if obs.project else None,
        "building_id": obs.building_id,
        "building_name": obs.building.name if obs.building else None,
        "floor_id": obs.floor_id,
        "floor_name": obs.floor.name if obs.floor else None,
        "exact_location": obs.exact_location,
        "obs_time": obs.obs_time,
        "obs_date": obs.obs_date,
        "contractor_user_id": obs.contractor_user_id,
        "contractor_user_ids": json.loads(obs.contractor_user_ids) if obs.contractor_user_ids else [],
        "contractor_name": obs.contractor.name if obs.contractor else None,
        "to_be_rectified_by": obs.to_be_rectified_by,
        "observer_name": obs.observer_name,
        "core_concern_id": obs.core_concern_id,
        "core_concern_name": obs.core_concern.name if obs.core_concern else None,
        "specific_concern_id": obs.specific_concern_id,
        "specific_concern_name": obs.specific_concern.name if obs.specific_concern else None,
        "specific_concern_text": obs.specific_concern_text,
        "possible_outcome": obs.possible_outcome,
        "severity": obs.severity,
        "probability": obs.probability,
        "risk_factor": obs.risk_factor,
        "risk_level": obs.risk_level,
        "root_cause_category_id": obs.root_cause_category_id,
        "root_cause_category_name": obs.root_cause_category.name if obs.root_cause_category else None,
        "root_cause_specific_id": obs.root_cause_specific_id,
        "root_cause_specific_name": obs.root_cause_specific.name if obs.root_cause_specific else None,
        "violation_id": obs.violation_id,
        "violation_name": obs.violation.name if obs.violation else None,
        "target_date_id": obs.target_date_id,
        "target_date_name": obs.target_date.name if obs.target_date else None,
        "target_date_actual": obs.target_date_actual,
        "eic_user_id": obs.eic_user_id,
        "eic_user_ids": json.loads(obs.eic_user_ids) if obs.eic_user_ids else [],
        "eic_user_name": obs.eic_user.name if obs.eic_user else None,
        "closed_at": obs.closed_at.isoformat() if obs.closed_at else None,
        "status": obs.status,
        "created_by": obs.created_by,
        "created_by_name": obs.creator.name if obs.creator else None,
        "created_at": obs.created_at.isoformat() if obs.created_at else None,
        "updated_at": obs.updated_at.isoformat() if obs.updated_at else None,
        "images": [
            {"id": img.id, "file_path": img.file_path, "file_name": img.file_name,
             "image_type": img.image_type,
             "uploaded_by": img.uploaded_by,
             "uploader_name": img.uploader.name if img.uploader else None,
             "uploader_role": img.uploader.role if img.uploader else None,
             "created_at": img.created_at.isoformat() if img.created_at else None}
            for img in images
        ],
        "creator_role": obs.creator.role if obs.creator else None,
        "comments": [
            {"id": c.id, "comment": c.comment, "user_id": c.user_id,
             "user_name": c.user.name if c.user else None,
             "user_role": c.user.role if c.user else None,
             "created_at": c.created_at.isoformat() if c.created_at else None}
            for c in comments
        ],
    }


class ObsCreate(BaseModel):
    project_id: int
    building_id: Optional[int] = None
    floor_id: Optional[int] = None
    exact_location: Optional[str] = None
    obs_time: Optional[str] = None
    obs_date: Optional[str] = None
    contractor_user_id: Optional[int] = None
    contractor_user_ids: Optional[List[int]] = None   # all selected contractor user IDs
    to_be_rectified_by: Optional[str] = None
    observer_name: Optional[str] = None
    core_concern_id: Optional[int] = None
    specific_concern_id: Optional[int] = None
    specific_concern_text: Optional[str] = None
    possible_outcome: Optional[str] = None
    severity: Optional[int] = None
    probability: Optional[int] = None
    root_cause_category_id: Optional[int] = None
    root_cause_specific_id: Optional[int] = None
    violation_id: Optional[int] = None
    target_date_id: Optional[int] = None
    target_date_actual: Optional[str] = None   # YYYY-MM-DD calendar date
    eic_user_id: Optional[int] = None
    eic_user_ids: Optional[List[int]] = None
    status: Optional[str] = None


class ObsUpdate(ObsCreate):
    status: Optional[str] = None


class CommentBody(BaseModel):
    comment: str


def _days_aging(target_str: Optional[str], closed_at, today: date) -> Optional[int]:
    """Return days between target date and closure (or today). Positive = overdue."""
    if not target_str:
        return None
    try:
        target = date.fromisoformat(target_str)
    except ValueError:
        return None
    if closed_at:
        end = closed_at.date() if isinstance(closed_at, datetime) else date.fromisoformat(str(closed_at)[:10])
    else:
        end = today
    return (end - target).days


def _aging_bucket(days: Optional[int]) -> str:
    if days is None:
        return "no_target"
    if days <= 0:
        return "on_time"
    if days <= 7:
        return "overdue_1_7"
    if days <= 30:
        return "overdue_8_30"
    return "overdue_30_plus"


def _matches_aging_filter(days: Optional[int], aging_filter: List[str]) -> bool:
    if not aging_filter:
        return True
    for f in aging_filter:
        if f == "no_target" and days is None:
            return True
        if f == "overdue" and days is not None and days > 0:
            return True
        if f == "due_soon" and days is not None and -7 <= days <= 0:
            return True
        if f == "on_time" and days is not None and days <= -7:
            return True
    return False


@router.get("/stats/summary")
def stats(
    project_id: List[int] = Query(default=[]),
    building_id: Optional[int] = Query(None),
    contractor_user_id: List[int] = Query(default=[]),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    core_concern_id: List[int] = Query(default=[]),
    risk_level: List[str] = Query(default=[]),
    aging: List[str] = Query(default=[]),
    db: Session = Depends(get_db),
    user: models.User = Depends(get_current_user),
):
    allowed = get_allowed_project_ids(user)
    today = date.today()

    def apply_filters(q):
        if allowed is not None:
            q = q.filter(models.Observation.project_id.in_(allowed))
        if project_id:
            q = q.filter(models.Observation.project_id.in_(project_id))
        if building_id:
            q = q.filter(models.Observation.building_id == building_id)
        if contractor_user_id:
            q = q.filter(models.Observation.contractor_user_id.in_(contractor_user_id))
        if date_from:
            q = q.filter(models.Observation.obs_date >= date_from)
        if date_to:
            q = q.filter(models.Observation.obs_date <= date_to)
        if core_concern_id:
            q = q.filter(models.Observation.core_concern_id.in_(core_concern_id))
        if risk_level:
            q = q.filter(models.Observation.risk_level.in_(risk_level))
        return q

    # Resolve aging filter → set of eligible observation IDs
    aging_ids: Optional[set] = None
    if aging:
        rows = (
            apply_filters(db.query(models.Observation.id, models.Observation.target_date_actual, models.Observation.closed_at))
            .filter(models.Observation.status != 'Draft')
            .all()
        )
        aging_ids = {r[0] for r in rows if _matches_aging_filter(_days_aging(r[1], r[2], today), aging)}

    def apply_all(q):
        q = apply_filters(q)
        if aging_ids is not None:
            q = q.filter(models.Observation.id.in_(aging_ids))
        return q

    q = apply_all(db.query(models.Observation)).filter(models.Observation.status != 'Draft')
    total = q.count()

    # Aging breakdown (always computed from base filters, ignoring aging filter for the donut itself)
    aging_rows = (
        apply_filters(db.query(models.Observation.target_date_actual, models.Observation.closed_at, models.Observation.status))
        .filter(models.Observation.status != 'Draft')
        .all()
    )
    aging_buckets: dict = {"positive_approach": 0, "no_target": 0, "on_time": 0, "overdue_1_7": 0, "overdue_8_30": 0, "overdue_30_plus": 0}
    for target_str, closed_at, status in aging_rows:
        if status == 'Positive Approach':
            aging_buckets["positive_approach"] += 1
        else:
            bucket = _aging_bucket(_days_aging(target_str, closed_at, today))
            aging_buckets[bucket] = aging_buckets.get(bucket, 0) + 1

    by_status = apply_all(
        db.query(models.Observation.status, func.count().label("count"))
    ).filter(models.Observation.status != 'Draft').group_by(models.Observation.status).all()

    by_risk = apply_all(
        db.query(models.Observation.risk_level, func.count().label("count"))
    ).filter(models.Observation.status != 'Draft').group_by(models.Observation.risk_level).all()

    recent = q.order_by(models.Observation.created_at.desc()).limit(4).all()

    # Monthly trend — use obs_date (YYYY-MM-DD string); substr gives YYYY-MM reliably
    month_expr = func.substr(models.Observation.obs_date, 1, 7)
    by_month = (
        apply_all(db.query(month_expr.label("month"), func.count().label("count")))
        .filter(models.Observation.status != 'Draft')
        .filter(models.Observation.obs_date.isnot(None))
        .filter(models.Observation.obs_date != "")
        .group_by(month_expr)
        .order_by(month_expr)
        .all()
    )

    # Monthly trend broken down by status
    by_month_status_rows = (
        apply_all(
            db.query(
                month_expr.label("month"),
                models.Observation.status,
                func.count().label("count"),
            )
        )
        .filter(models.Observation.status != 'Draft')
        .filter(models.Observation.obs_date.isnot(None))
        .filter(models.Observation.obs_date != "")
        .group_by(month_expr, models.Observation.status)
        .order_by(month_expr)
        .all()
    )
    # Pivot into { month, Open, Pending, Under Review, Partially Closed, Closed, Positive Approach }
    month_status_map: dict = {}
    for m, s, c in by_month_status_rows:
        if not m:
            continue
        if m not in month_status_map:
            month_status_map[m] = {"month": m, "Open": 0, "Overdue": 0, "Under Review": 0, "Partially Closed": 0, "Closed": 0, "Positive Approach": 0}
        if s in month_status_map[m]:
            month_status_map[m][s] = c

    by_month_status = sorted(month_status_map.values(), key=lambda x: x["month"])

    return {
        "total": total,
        "byStatus": [{"status": s, "count": c} for s, c in by_status],
        "byRisk": [{"risk_level": r, "count": c} for r, c in by_risk if r],
        "byMonth": [{"month": m, "count": c} for m, c in by_month if m],
        "byMonthStatus": by_month_status,
        "byAging": aging_buckets,
        "recent": [
            {
                "observation_id": o.observation_id,
                "status": o.status,
                "risk_level": o.risk_level,
                "obs_date": o.obs_date,
                "created_at": o.created_at.isoformat() if o.created_at else None,
                "project_name": o.project.name if o.project else None,
                "core_concern_name": o.core_concern.name if o.core_concern else None,
            }
            for o in recent
        ],
    }


def to_fy_quarter(year_str, month_str) -> str:
    if not year_str or not month_str:
        return ""
    year = int(year_str)
    month = int(month_str)
    if month >= 4:
        fy_start = year
        if month <= 6:
            q = "Q-1"
        elif month <= 9:
            q = "Q-2"
        else:
            q = "Q-3"
    else:
        fy_start = year - 1
        q = "Q-4"
    fy_end = fy_start + 1
    return f"{q} ({str(fy_start)[-2:]}-{str(fy_end)[-2:]})"


@router.get("/stats/summary-details")
def stats_details(
    project_id: List[int] = Query(default=[]),
    building_id: Optional[int] = Query(None),
    contractor_user_id: List[int] = Query(default=[]),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    user: models.User = Depends(get_current_user),
):
    allowed = get_allowed_project_ids(user)
    conditions = []
    if allowed is not None:
        conditions.append(models.Observation.project_id.in_(allowed))
    if project_id:
        conditions.append(models.Observation.project_id.in_(project_id))
    if building_id:
        conditions.append(models.Observation.building_id == building_id)
    if contractor_user_id:
        conditions.append(models.Observation.contractor_user_id.in_(contractor_user_id))
    if date_from:
        conditions.append(models.Observation.obs_date >= date_from)
    if date_to:
        conditions.append(models.Observation.obs_date <= date_to)

    project_summary_rows = db.query(
        models.Project.id.label("project_id"),
        models.Project.name.label("project_name"),
        func.count().label("total"),
        func.sum(case((models.Observation.status == "Closed", 1), else_=0)).label("closed"),
        func.sum(case((models.Observation.risk_level == "High", 1), else_=0)).label("high_risk"),
        func.sum(case((models.Observation.risk_level == "Medium", 1), else_=0)).label("medium_risk"),
        func.sum(case((models.Observation.risk_level == "Low", 1), else_=0)).label("low_risk"),
    ).join(models.Project, models.Observation.project_id == models.Project.id)
    if conditions:
        project_summary_rows = project_summary_rows.filter(*conditions)
    project_summary_rows = project_summary_rows.group_by(models.Project.id, models.Project.name).order_by(func.count().desc()).all()

    ContractorUser = models.User
    contractor_summary_rows = db.query(
        models.Observation.contractor_user_id.label("contractor_id"),
        ContractorUser.name.label("contractor_name"),
        func.count().label("total"),
        func.sum(case((models.Observation.status == "Closed", 1), else_=0)).label("closed"),
        func.sum(case((models.Observation.risk_level == "High", 1), else_=0)).label("high_risk"),
        func.sum(case((models.Observation.risk_level == "Medium", 1), else_=0)).label("medium_risk"),
        func.sum(case((models.Observation.risk_level == "Low", 1), else_=0)).label("low_risk"),
    ).join(ContractorUser, models.Observation.contractor_user_id == ContractorUser.id
    ).filter(models.Observation.contractor_user_id != None)
    if conditions:
        contractor_summary_rows = contractor_summary_rows.filter(*conditions)
    contractor_summary_rows = contractor_summary_rows.group_by(
        models.Observation.contractor_user_id,
        ContractorUser.name,
    ).order_by(func.count().desc()).all()

    from collections import defaultdict

    category_month_rows = db.query(
        models.CoreConcern.name.label("category"),
        func.substr(models.Observation.obs_date, 1, 4).label("year"),
        func.substr(models.Observation.obs_date, 6, 2).label("month"),
        func.sum(case((models.Observation.risk_level == "Low", 1), else_=0)).label("low"),
        func.sum(case((models.Observation.risk_level == "Medium", 1), else_=0)).label("medium"),
        func.sum(case((models.Observation.risk_level == "High", 1), else_=0)).label("high"),
        func.count().label("total"),
    ).join(models.CoreConcern, models.Observation.core_concern_id == models.CoreConcern.id)
    if conditions:
        category_month_rows = category_month_rows.filter(*conditions)
    category_month_rows = category_month_rows.group_by(
        models.CoreConcern.name,
        func.substr(models.Observation.obs_date, 1, 4),
        func.substr(models.Observation.obs_date, 6, 2),
    ).order_by(
        func.substr(models.Observation.obs_date, 1, 4),
        func.substr(models.Observation.obs_date, 6, 2),
        models.CoreConcern.name,
    ).all()

    # Aggregate by Indian FY quarter
    agg: dict = defaultdict(lambda: {"low": 0, "medium": 0, "high": 0, "total": 0})
    for row in category_month_rows:
        key = (row.category, to_fy_quarter(row.year, row.month))
        agg[key]["low"] += row.low or 0
        agg[key]["medium"] += row.medium or 0
        agg[key]["high"] += row.high or 0
        agg[key]["total"] += row.total or 0

    category_scores = [
        {
            "category": cat,
            "quarter": quarter,
            "low": d["low"],
            "medium": d["medium"],
            "high": d["high"],
            "total": d["total"],
            "score": round(((d["low"] + d["medium"]) / d["total"]) * 100) if d["total"] else 0,
        }
        for (cat, quarter), d in agg.items()
    ]

    observer_rows = db.query(
        func.coalesce(models.Observation.observer_name, models.User.name).label("observer_name"),
        func.count().label("count"),
    ).join(models.User, models.Observation.created_by == models.User.id)
    if conditions:
        observer_rows = observer_rows.filter(*conditions)
    observer_rows = observer_rows.group_by(
        func.coalesce(models.Observation.observer_name, models.User.name)
    ).order_by(func.count().desc()).limit(10).all()

    def build_summary(rows, label_fields):
        return [
            {
                **{field: getattr(row, field) for field in label_fields},
                "total": row.total,
                "closed": row.closed,
                "high_risk": row.high_risk,
                "medium_risk": row.medium_risk,
                "low_risk": row.low_risk,
                "compliance_score": round(((row.low_risk or 0) + (row.medium_risk or 0)) / (row.total or 1) * 100),
            }
            for row in rows
        ]

    return {
        "projectSummary": build_summary(project_summary_rows, ["project_id", "project_name"]),
        "contractorSummary": build_summary(contractor_summary_rows, ["contractor_id", "contractor_name"]),
        "categoryQuarterScores": category_scores,
        "topObservers": [
            {"observer_name": row.observer_name or "Unknown", "count": row.count}
            for row in observer_rows
        ],
        "formula": "score = round((low + medium) / total * 100)",
    }


@router.get("/")
def list_observations(
    project_id: List[int] = Query(default=[]),
    status: List[str] = Query(default=[]),
    contractor_user_id: List[int] = Query(default=[]),
    risk_level: List[str] = Query(default=[]),
    core_concern_id: List[int] = Query(default=[]),
    specific_concern_id: List[int] = Query(default=[]),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=10000),
    db: Session = Depends(get_db),
    user: models.User = Depends(get_current_user),
):
    allowed = get_allowed_project_ids(user)
    q = db.query(models.Observation)
    if allowed is not None:
        if not allowed:
            return {"observations": [], "total": 0, "pages": 0}
        q = q.filter(models.Observation.project_id.in_(allowed))
    if project_id:
        q = q.filter(models.Observation.project_id.in_(project_id))
    if status:
        q = q.filter(models.Observation.status.in_(status))
    if contractor_user_id:
        q = q.filter(models.Observation.contractor_user_id.in_(contractor_user_id))
    if risk_level:
        q = q.filter(models.Observation.risk_level.in_(risk_level))
    if core_concern_id:
        q = q.filter(models.Observation.core_concern_id.in_(core_concern_id))
    if specific_concern_id:
        q = q.filter(models.Observation.specific_concern_id.in_(specific_concern_id))
    if date_from:
        q = q.filter(models.Observation.obs_date >= date_from)
    if date_to:
        q = q.filter(models.Observation.obs_date <= date_to)

    # Drafts are private — visible to creator or admins
    if user.role not in ("SuperAdmin", "Admin"):
        q = q.filter(
            (models.Observation.status != 'Draft') | (models.Observation.created_by == user.id)
        )

    total = q.count()
    obs_list = q.order_by(
        nullslast(models.Observation.created_at.desc()),
        models.Observation.obs_date.desc(),
    ).offset((page - 1) * limit).limit(limit).all()

    return {
        "observations": [obs_to_dict(o, db) for o in obs_list],
        "total": total,
        "pages": (total + limit - 1) // limit,
    }


@router.get("/stats/she-report")
def she_report_stats(
    project_id: List[int] = Query(default=[]),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    user: models.User = Depends(get_current_user),
):
    """Comprehensive SHE performance data for the 7-page report PDF."""
    from collections import defaultdict

    allowed = get_allowed_project_ids(user)
    today = date.today()

    conditions = [models.Observation.status != "Draft"]
    if allowed is not None:
        conditions.append(models.Observation.project_id.in_(allowed))
    if project_id:
        conditions.append(models.Observation.project_id.in_(project_id))
    if date_from:
        conditions.append(models.Observation.obs_date >= date_from)
    if date_to:
        conditions.append(models.Observation.obs_date <= date_to)

    obs_list = db.query(models.Observation).filter(*conditions).all()

    # ── Per-project rectification stats ──────────────────────────────────
    proj_stats: dict = defaultdict(lambda: {
        "project_name": "", "raised": 0, "rectified": 0, "not_rectified": 0,
        "timely": 0, "delayed": 0, "total_delay_days": 0,
    })

    for obs in obs_list:
        pname = obs.project.name if obs.project else "Unknown"
        ps = proj_stats[obs.project_id or 0]
        ps["project_name"] = pname
        ps["raised"] += 1
        if obs.status == "Closed":
            ps["rectified"] += 1
            if obs.closed_at and obs.target_date_actual:
                try:
                    target = date.fromisoformat(obs.target_date_actual)
                    closed_d = obs.closed_at.date() if hasattr(obs.closed_at, "date") else obs.closed_at
                    delay = (closed_d - target).days
                    if delay <= 0:
                        ps["timely"] += 1
                    else:
                        ps["delayed"] += 1
                        ps["total_delay_days"] += delay
                except (ValueError, AttributeError):
                    ps["timely"] += 1
            else:
                ps["timely"] += 1
        else:
            ps["not_rectified"] += 1
            if obs.target_date_actual:
                try:
                    target = date.fromisoformat(obs.target_date_actual)
                    if today > target:
                        ps["total_delay_days"] += (today - target).days
                except ValueError:
                    pass

    project_rectification = sorted([
        {
            "project_name": ps["project_name"],
            "raised": ps["raised"],
            "rectified": ps["rectified"],
            "not_rectified": ps["not_rectified"],
            "timely": ps["timely"],
            "delayed": ps["delayed"],
            "total_delay_days": ps["total_delay_days"],
            "avg_delay": round(ps["total_delay_days"] / ps["raised"], 2) if ps["raised"] else 0,
        }
        for ps in proj_stats.values()
    ], key=lambda x: -x["raised"])

    # ── Consequence distribution (possible_outcome field) ─────────────────
    consequence_counts: dict = defaultdict(int)
    for obs in obs_list:
        if obs.possible_outcome:
            consequence_counts[obs.possible_outcome.strip()] += 1
    total_cons = sum(consequence_counts.values()) or 1
    consequence_distribution = sorted([
        {"name": k, "count": v, "pct": round(v / total_cons * 100)}
        for k, v in consequence_counts.items()
    ], key=lambda x: -x["count"])

    # ── Root cause distribution ───────────────────────────────────────────
    root_cause_counts: dict = defaultdict(int)
    for obs in obs_list:
        if obs.root_cause_category:
            root_cause_counts[obs.root_cause_category.name] += 1
    total_rc = sum(root_cause_counts.values()) or 1
    root_cause_distribution = sorted([
        {"name": k, "count": v, "pct": round(v / total_rc * 100)}
        for k, v in root_cause_counts.items()
    ], key=lambda x: -x["count"])

    # ── Violation area distribution (core concern) ────────────────────────
    violation_area_counts: dict = defaultdict(int)
    for obs in obs_list:
        if obs.core_concern:
            violation_area_counts[obs.core_concern.name] += 1
    total_va = sum(violation_area_counts.values()) or 1
    violation_area_distribution = sorted([
        {"name": k, "count": v, "pct": round(v / total_va * 100)}
        for k, v in violation_area_counts.items()
    ], key=lambda x: -x["count"])

    # ── Per-project risk analysis ─────────────────────────────────────────
    proj_risk: dict = defaultdict(lambda: {"project_name": "", "total": 0, "high": 0, "medium": 0, "low": 0})
    for obs in obs_list:
        pname = obs.project.name if obs.project else "Unknown"
        pr = proj_risk[obs.project_id or 0]
        pr["project_name"] = pname
        pr["total"] += 1
        if obs.risk_level == "High":
            pr["high"] += 1
        elif obs.risk_level == "Medium":
            pr["medium"] += 1
        elif obs.risk_level == "Low":
            pr["low"] += 1

    project_risk_analysis = sorted(proj_risk.values(), key=lambda x: -x["total"])

    total_raised = sum(p["raised"] for p in project_rectification)
    total_delay = sum(p["total_delay_days"] for p in project_rectification)

    return {
        "projectRectification": project_rectification,
        "consequenceDistribution": consequence_distribution,
        "rootCauseDistribution": root_cause_distribution,
        "violationAreaDistribution": violation_area_distribution,
        "projectRiskAnalysis": project_risk_analysis,
        "avgDelayOverall": round(total_delay / total_raised, 2) if total_raised else 0,
        "totalObservations": len(obs_list),
    }


@router.get("/report")
def get_report_data(
    project_id: List[int] = Query(default=[]),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    status: List[str] = Query(default=[]),
    contractor_user_id: List[int] = Query(default=[]),
    risk_level: List[str] = Query(default=[]),
    db: Session = Depends(get_db),
    user: models.User = Depends(get_current_user),
):
    allowed = get_allowed_project_ids(user)
    q = db.query(models.Observation)
    if allowed is not None:
        q = q.filter(models.Observation.project_id.in_(allowed))
    if project_id:
        q = q.filter(models.Observation.project_id.in_(project_id))
    if date_from:
        q = q.filter(models.Observation.obs_date >= date_from)
    if date_to:
        q = q.filter(models.Observation.obs_date <= date_to)
    q = q.filter(models.Observation.status != 'Draft')
    if status:
        q = q.filter(models.Observation.status.in_(status))
    if contractor_user_id:
        q = q.filter(models.Observation.contractor_user_id.in_(contractor_user_id))
    if risk_level:
        q = q.filter(models.Observation.risk_level.in_(risk_level))
    obs_list = (
        q.order_by(models.Observation.obs_date.asc(), models.Observation.observation_id.asc())
        .limit(200)
        .all()
    )
    return [obs_to_dict(o, db) for o in obs_list]


@router.get("/{obs_id}")
def get_observation(obs_id: str, db: Session = Depends(get_db), user: models.User = Depends(get_current_user)):
    # obs_id can be numeric id or observation_id string
    if obs_id.isdigit():
        obs = db.query(models.Observation).filter(models.Observation.id == int(obs_id)).first()
    else:
        obs = db.query(models.Observation).filter(models.Observation.observation_id == obs_id).first()
    if not obs:
        raise HTTPException(404, "Observation not found")
    if obs.status == 'Draft' and obs.created_by != user.id:
        raise HTTPException(404, "Observation not found")
    _assert_obs_access(obs, user)
    return obs_to_dict(obs, db)


@router.post("/", status_code=201)
def create_observation(body: ObsCreate, db: Session = Depends(get_db), user: models.User = Depends(get_current_user)):
    if user.role == "Contractor":
        raise HTTPException(status_code=403, detail="Your role cannot create observations")
    factor, level = calc_risk(body.severity or 1, body.probability or 1)
    obs_id = generate_obs_id(db, body.project_id)

    obs = models.Observation(
        observation_id=obs_id,
        project_id=body.project_id,
        building_id=body.building_id,
        floor_id=body.floor_id,
        exact_location=body.exact_location,
        obs_time=body.obs_time,
        obs_date=body.obs_date,
        contractor_user_id=body.contractor_user_id,
        contractor_user_ids=json.dumps(body.contractor_user_ids) if body.contractor_user_ids else None,
        to_be_rectified_by=body.to_be_rectified_by,
        observer_name=body.observer_name,
        core_concern_id=body.core_concern_id,
        specific_concern_id=body.specific_concern_id,
        specific_concern_text=body.specific_concern_text,
        possible_outcome=body.possible_outcome,
        severity=body.severity,
        probability=body.probability,
        risk_factor=factor,
        risk_level=level,
        root_cause_category_id=body.root_cause_category_id,
        root_cause_specific_id=body.root_cause_specific_id,
        violation_id=body.violation_id,
        target_date_id=body.target_date_id,
        target_date_actual=body.target_date_actual,
        eic_user_id=body.eic_user_ids[0] if body.eic_user_ids else body.eic_user_id,
        eic_user_ids=json.dumps(body.eic_user_ids) if body.eic_user_ids else None,
        status=body.status or "Open",
        created_by=user.id,
    )
    db.add(obs)
    db.commit()
    db.refresh(obs)

    # In-app notification and email — skipped for drafts
    if obs.status == 'Draft':
        return {"id": obs.id, "observation_id": obs.observation_id, "risk_factor": factor, "risk_level": level}

    # In-app notifications to all assigned contractors
    notif_ids = body.contractor_user_ids or ([body.contractor_user_id] if body.contractor_user_id else [])
    project_name = obs.project.name if obs.project else ""
    for uid in notif_ids:
        db.add(models.Notification(
            user_id=uid,
            observation_id=obs.id,
            obs_ref=obs.observation_id,
            message=f"New observation {obs.observation_id} assigned to you on project '{project_name}'.",
            is_read=False,
        ))
        db.commit()

    # Email notification
    threading.Thread(target=_fire_email_async, args=(obs.id,), daemon=True).start()

    return {"id": obs.id, "observation_id": obs.observation_id, "risk_factor": factor, "risk_level": level}


def _send_obs_email(obs: models.Observation, db: Session, event: str = "new"):
    smtp = db.query(models.SmtpSettings).first()
    if not smtp:
        logger.warning("Email skipped for %s: no SMTP settings row in DB", obs.observation_id)
        return
    if not smtp.enabled:
        logger.warning("Email skipped for %s: SMTP is disabled (enabled=False)", obs.observation_id)
        return

    # TO: all assigned contractors
    to_emails: List[str] = []
    contractor_names: List[str] = []
    contractor_ids = json.loads(obs.contractor_user_ids) if obs.contractor_user_ids else (
        [obs.contractor_user_id] if obs.contractor_user_id else []
    )
    for cid in contractor_ids:
        u = db.query(models.User).filter(models.User.id == cid).first()
        if u:
            if u.email and u.email not in to_emails:
                to_emails.append(u.email)
            if u.name and u.name not in contractor_names:
                contractor_names.append(u.name)

    # Observation-specific EIC users
    eic_names: List[str] = []
    eic_ids = json.loads(obs.eic_user_ids) if obs.eic_user_ids else (
        [obs.eic_user_id] if obs.eic_user_id else []
    )
    eic_cc: List[str] = []
    for eid in eic_ids:
        u = db.query(models.User).filter(models.User.id == eid).first()
        if u:
            if u.email and u.email not in to_emails and u.email not in eic_cc:
                eic_cc.append(u.email)
            if u.name and u.name not in eic_names:
                eic_names.append(u.name)

    # CC: project-scoped PIC/EIC/HO/PSO/Observer + Admin/SuperAdmin (project-filtered or global)
    CC_PROJECT_ROLES = {"PIC", "EIC", "HO", "PSO", "Observer"}
    cc_emails: List[str] = list(eic_cc)
    project_users = (
        db.query(models.User)
        .join(models.UserProject, models.UserProject.user_id == models.User.id)
        .filter(
            models.UserProject.project_id == obs.project_id,
            models.User.role.in_(CC_PROJECT_ROLES),
        )
        .all()
    )
    # Admin/SuperAdmin: include if they have no project assignments (All Projects)
    # or if they are assigned to this specific project
    all_admins = db.query(models.User).filter(models.User.role.in_(["Admin", "SuperAdmin"])).all()
    admin_project_ids_map = {}
    for u in all_admins:
        admin_project_ids_map[u.id] = {up.project_id for up in u.user_projects}
    admin_users = [
        u for u in all_admins
        if not admin_project_ids_map[u.id]  # empty = All Projects
        or obs.project_id in admin_project_ids_map[u.id]
    ]
    for u in project_users + admin_users:
        if u.email and u.email not in cc_emails and u.email not in to_emails:
            cc_emails.append(u.email)

    if not to_emails and not cc_emails:
        logger.warning("Email skipped for %s: no recipients found (no contractors/EICs/project users)", obs.observation_id)
        return

    obs_data = {
        "observation_id": obs.observation_id,
        "project_name": obs.project.name if obs.project else "",
        "obs_date": obs.obs_date or "",
        "obs_time": obs.obs_time or "",
        "building_name": obs.building.name if obs.building else "",
        "floor_name": obs.floor.name if obs.floor else "",
        "exact_location": obs.exact_location or "",
        "observer_name": obs.observer_name or "",
        "created_by_name": obs.creator.name if obs.creator else "",
        "contractor_names": contractor_names,
        "eic_names": eic_names,
        "to_be_rectified_by": obs.to_be_rectified_by or "",
        "core_concern_name": obs.core_concern.name if obs.core_concern else "",
        "specific_concern_name": obs.specific_concern.name if obs.specific_concern else "",
        "specific_concern_text": obs.specific_concern_text or "",
        "possible_outcome": obs.possible_outcome or "",
        "severity": obs.severity,
        "probability": obs.probability,
        "risk_factor": obs.risk_factor,
        "risk_level": obs.risk_level or "",
        "root_cause_category_name": obs.root_cause_category.name if obs.root_cause_category else "",
        "root_cause_specific_name": obs.root_cause_specific.name if obs.root_cause_specific else "",
        "violation_name": obs.violation.name if obs.violation else "",
        "target_date_name": obs.target_date.name if obs.target_date else "",
        "target_date_actual": obs.target_date_actual or "",
        "status": obs.status,
        "event": event,
    }
    subject, html = build_observation_email(obs_data)
    send_observation_email(smtp, to_emails, cc_emails, subject, html)


def _fire_email_async(obs_id: int, event: str = "new"):
    """Send observation email in a background thread so it doesn't block the HTTP response."""
    db = SessionLocal()
    try:
        obs = db.query(models.Observation).filter(models.Observation.id == obs_id).first()
        if obs:
            _send_obs_email(obs, db, event=event)
    except Exception as exc:
        logger.error("Async email failed for observation %s: %s", obs_id, exc)
    finally:
        db.close()


@router.put("/{obs_id}")
def update_observation(obs_id: int, body: ObsUpdate, db: Session = Depends(get_db), user: models.User = Depends(get_current_user)):
    obs = db.query(models.Observation).filter(models.Observation.id == obs_id).first()
    if not obs:
        raise HTTPException(404, "Observation not found")
    _assert_obs_access(obs, user, write_roles=['SuperAdmin', 'Admin', 'HO', 'PSO', 'Observer'])

    was_draft = obs.status == 'Draft'
    prev_status = obs.status
    factor, level = calc_risk(body.severity or obs.severity or 1, body.probability or obs.probability or 1)

    body_dict = body.model_dump(exclude_unset=True)
    # Serialize list fields → JSON string before setting on model
    if 'contractor_user_ids' in body_dict:
        ids = body_dict.pop('contractor_user_ids')
        obs.contractor_user_ids = json.dumps(ids) if ids else None
    if 'eic_user_ids' in body_dict:
        eids = body_dict.pop('eic_user_ids')
        obs.eic_user_ids = json.dumps(eids) if eids else None
        obs.eic_user_id = eids[0] if eids else body_dict.pop('eic_user_id', obs.eic_user_id)
    for field, val in body_dict.items():
        setattr(obs, field, val)

    obs.risk_factor = factor
    obs.risk_level = level
    obs.updated_at = datetime.now()
    if prev_status != 'Closed' and obs.status == 'Closed' and obs.closed_at is None:
        obs.closed_at = datetime.now()
    db.commit()
    db.refresh(obs)

    # Send notifications when a draft is being submitted (converted to non-Draft)
    if was_draft and obs.status != 'Draft':
        all_cids = json.loads(obs.contractor_user_ids) if obs.contractor_user_ids else (
            [obs.contractor_user_id] if obs.contractor_user_id else []
        )
        if all_cids:
            project_name = obs.project.name if obs.project else ""
            for cid in all_cids:
                db.add(models.Notification(
                    user_id=cid, observation_id=obs.id, obs_ref=obs.observation_id,
                    message=f"New observation {obs.observation_id} assigned to you on project '{project_name}'.",
                    is_read=False,
                ))
            db.commit()
        threading.Thread(target=_fire_email_async, args=(obs.id,), daemon=True).start()

    return {"success": True, "risk_factor": factor, "risk_level": level}


class StatusBody(BaseModel):
    status: str


@router.patch("/{obs_id}/status")
def update_status(obs_id: int, body: StatusBody, db: Session = Depends(get_db), user: models.User = Depends(get_current_user)):
    obs = db.query(models.Observation).filter(models.Observation.id == obs_id).first()
    if not obs:
        raise HTTPException(404, "Observation not found")
    _assert_obs_access(obs, user, write_roles=['SuperAdmin', 'Admin', 'HO', 'Observer', 'Contractor'])
    prev_status = obs.status
    obs.status = body.status
    obs.updated_at = datetime.now()
    if prev_status != 'Closed' and body.status == 'Closed' and obs.closed_at is None:
        obs.closed_at = datetime.now()
    db.commit()

    if prev_status != body.status:
        threading.Thread(target=_fire_email_async, args=(obs.id, "status_change"), daemon=True).start()

    return {"success": True, "status": body.status}


@router.post("/{obs_id}/comments", status_code=201)
def add_comment(obs_id: int, body: CommentBody, db: Session = Depends(get_db), user: models.User = Depends(get_current_user)):
    obs = db.query(models.Observation).filter(models.Observation.id == obs_id).first()
    if not obs:
        raise HTTPException(404)
    _assert_obs_access(obs, user)  # all authenticated users with project access may comment
    c = models.ObservationComment(observation_id=obs_id, user_id=user.id, comment=body.comment)
    db.add(c)
    obs.updated_at = datetime.now()
    db.commit()
    db.refresh(c)
    return {"id": c.id}


@router.delete("/{obs_id}")
def delete_observation(obs_id: int, db: Session = Depends(get_db), user: models.User = Depends(get_current_user)):
    obs = db.query(models.Observation).filter(models.Observation.id == obs_id).first()
    if not obs:
        raise HTTPException(404)
    is_super_admin = user.role == "SuperAdmin"
    is_own_draft = obs.status == "Draft" and obs.created_by == user.id
    if not is_super_admin and not is_own_draft:
        raise HTTPException(403, "Only SuperAdmin can delete observations, or the creator can delete their own draft")
    db.delete(obs)
    db.commit()
    return {"success": True}


# ── Bulk import ──────────────────────────────────────────────────────────────

def _clean(val) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip()
    return None if s.lower() in ('none', 'nil', '', 'n/a') else s


def _fmt_date(val) -> Optional[str]:
    if val is None:
        return None
    if hasattr(val, 'strftime'):
        return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    if not s or s.lower() in ('none', 'nil'):
        return None
    for fmt in ('%d-%m-%Y', '%Y-%m-%d', '%d/%m/%Y'):
        try:
            return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
        except ValueError:
            pass
    return None


def _fmt_time(val) -> Optional[str]:
    if val is None:
        return None
    if hasattr(val, 'strftime'):
        return val.strftime('%H:%M')
    s = str(val).strip()
    return s[:5] if s and s.lower() != 'none' else None


def _safe_int(val) -> Optional[int]:
    try:
        v = int(float(str(val)))
        return v if v > 0 else None
    except (TypeError, ValueError):
        return None


def _find_project(db: Session, name: str) -> models.Project:
    """Case-insensitive exact match → fuzzy fallback (≥0.7) → auto-create."""
    project = db.query(models.Project).filter(
        models.Project.name.ilike(name)
    ).first()
    if project:
        return project
    all_projects = db.query(models.Project).all()
    name_map = {p.name.lower(): p for p in all_projects}
    matches = get_close_matches(name.lower(), name_map.keys(), n=1, cutoff=0.7)
    if matches:
        return name_map[matches[0]]
    new_project = models.Project(name=name.strip().upper())
    db.add(new_project)
    db.flush()
    return new_project


def _resolve_or_create_building(db: Session, name: str, project_id: int) -> models.Building:
    obj = db.query(models.Building).filter(
        models.Building.project_id == project_id,
        models.Building.name.ilike(name),
    ).first()
    if not obj:
        obj = models.Building(name=name, project_id=project_id)
        db.add(obj)
        db.flush()
    return obj


def _resolve_or_create_floor(db: Session, name: str, building_id: int) -> models.Floor:
    obj = db.query(models.Floor).filter(
        models.Floor.building_id == building_id,
        models.Floor.name.ilike(name),
    ).first()
    if not obj:
        obj = models.Floor(name=name, building_id=building_id)
        db.add(obj)
        db.flush()
    return obj


@router.post("/bulk-import")
async def bulk_import_observations(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
):
    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    except Exception:
        raise HTTPException(400, "Invalid Excel file")

    if 'Observations' not in wb.sheetnames:
        raise HTTPException(400, "Sheet named 'Observations' not found")

    ws = wb['Observations']
    rows = list(ws.iter_rows(values_only=True))

    # Find the actual header row (contains "Observation ID" in col 0)
    header_idx = None
    for i, row in enumerate(rows):
        if row and str(row[0]).strip() == 'Observation ID':
            header_idx = i
            break
    if header_idx is None:
        raise HTTPException(400, "Could not find 'Observation ID' header row")

    created, skipped, errors = [], [], []

    for row in rows[header_idx + 1:]:
        if not row or not row[0]:
            continue
        obs_id = str(row[0]).strip()
        if not obs_id:
            continue

        # Skip duplicates
        if db.query(models.Observation).filter(models.Observation.observation_id == obs_id).first():
            skipped.append(f"{obs_id} — already exists")
            continue

        try:
            obs_date        = _fmt_date(row[1])
            obs_time        = _fmt_time(row[2])
            project_name    = _clean(row[3])
            building_name   = _clean(row[4])
            floor_name      = _clean(row[5])
            exact_location  = _clean(row[6])
            observer_name   = _clean(row[7])
            contractor_name = _clean(row[8])
            rectified_by    = _clean(row[9])
            core_concern_name      = _clean(row[11])
            specific_concern_name  = _clean(row[12])
            specific_concern_text  = _clean(row[13])
            possible_outcome       = _clean(row[14])
            severity         = _safe_int(row[15])
            probability      = _safe_int(row[16])
            risk_level_raw   = _clean(row[18])
            root_cat_name    = _clean(row[19])
            root_spec_name   = _clean(row[20])
            violation_name   = _clean(row[21])
            closing_date     = _fmt_date(row[22])
            status           = _clean(row[23]) or 'Open'

            # Project (required)
            if not project_name:
                errors.append(f"{obs_id}: missing project name")
                continue
            project = _find_project(db, project_name)

            # Building / Floor — auto-create if absent
            building = _resolve_or_create_building(db, building_name, project.id) if building_name else None
            floor    = _resolve_or_create_floor(db, floor_name, building.id) if (floor_name and building) else None

            # Contractor user — prefer project-scoped match
            contractor_user = None
            if contractor_name:
                contractor_user = (
                    db.query(models.User)
                    .join(models.UserProject, models.UserProject.user_id == models.User.id)
                    .filter(
                        models.User.role == 'Contractor',
                        models.User.name.ilike(contractor_name),
                        models.UserProject.project_id == project.id,
                    ).first()
                )
                if not contractor_user:
                    contractor_user = db.query(models.User).filter(
                        models.User.role == 'Contractor',
                        models.User.name.ilike(contractor_name),
                    ).first()

            # Core concern → specific concern
            core_concern = None
            if core_concern_name:
                core_concern = db.query(models.CoreConcern).filter(
                    models.CoreConcern.name.ilike(core_concern_name)
                ).first()

            specific_concern = None
            if specific_concern_name and core_concern:
                specific_concern = db.query(models.SpecificConcern).filter(
                    models.SpecificConcern.core_concern_id == core_concern.id,
                    models.SpecificConcern.name.ilike(specific_concern_name),
                ).first()

            # Root cause
            root_cat = None
            if root_cat_name:
                root_cat = db.query(models.RootCauseCategory).filter(
                    models.RootCauseCategory.name.ilike(root_cat_name)
                ).first()

            root_spec = None
            if root_spec_name and root_cat:
                root_spec = db.query(models.RootCauseSpecific).filter(
                    models.RootCauseSpecific.root_cause_category_id == root_cat.id,
                    models.RootCauseSpecific.name.ilike(root_spec_name),
                ).first()

            # Violation
            violation = None
            if violation_name:
                violation = db.query(models.Violation).filter(
                    models.Violation.name.ilike(violation_name)
                ).first()

            # Risk
            risk_factor = (severity * probability) if (severity and probability) else None
            risk_level  = None if (not risk_level_raw or risk_level_raw.upper() == 'NIL') else risk_level_raw

            # closed_at
            closed_at = None
            if status == 'Closed' and closing_date:
                try:
                    closed_at = datetime.strptime(closing_date, '%Y-%m-%d')
                except ValueError:
                    pass

            obs = models.Observation(
                observation_id=obs_id,
                project_id=project.id,
                building_id=building.id if building else None,
                floor_id=floor.id if floor else None,
                exact_location=exact_location,
                obs_date=obs_date,
                obs_time=obs_time,
                contractor_user_id=contractor_user.id if contractor_user else None,
                contractor_user_ids=json.dumps([contractor_user.id]) if contractor_user else None,
                to_be_rectified_by=rectified_by,
                observer_name=observer_name,
                core_concern_id=core_concern.id if core_concern else None,
                specific_concern_id=specific_concern.id if specific_concern else None,
                specific_concern_text=specific_concern_text,
                possible_outcome=possible_outcome,
                severity=severity,
                probability=probability,
                risk_factor=risk_factor,
                risk_level=risk_level,
                root_cause_category_id=root_cat.id if root_cat else None,
                root_cause_specific_id=root_spec.id if root_spec else None,
                violation_id=violation.id if violation else None,
                target_date_actual=closing_date,
                status=status,
                closed_at=closed_at,
                created_by=current_user.id,
            )
            db.add(obs)
            created.append(f"{obs_id} → {project.name}")

        except Exception as exc:
            db.rollback()
            errors.append(f"{obs_id}: {exc}")

    db.commit()
    return {
        "created_count": len(created),
        "skipped_count": len(skipped),
        "error_count": len(errors),
        "created": created,
        "skipped": skipped,
        "errors": errors,
    }
