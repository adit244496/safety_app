"""
Generates the Safety Observation data-migration Excel template.
Run: python generate_migration_template.py
Output: Safety_Observation_Migration_Template.xlsx  (in the same directory)
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── Reference data (mirrors seed.py) ─────────────────────────────────────────

CATEGORIES = [
    "General safety", "Housekeeping", "PPE and Work apparels", "Work at Height",
    "Electrical", "Welding cutting grinding", "Vehicles", "Earth Moving Equipment",
    "Lifting Appliances and Gear", "Fire Prevention / Protection", "First Aid and Medicals",
    "Welfare facilities and Labour accommodation", "Environment Aspects", "Piling Work",
    "Excavation & Confined space job", "Hand tools and Power tools",
]

CORE_CONCERNS = [
    ("General safety", "Personal Protective equipment"),
    ("Work at Height", "Floor edge (Work at height)"),
    ("Work at Height", "Floor opening (Work at height)"),
    ("General safety", "Machinery"),
    ("Work at Height", "Scaffold (Work at height)"),
    ("Work at Height", "Ladder (Work at height)"),
    ("Work at Height", "Work platform (Work at height)"),
    ("Electrical", "Electrical"),
    ("Hand tools and Power tools", "Hand Tools Manual"),
    ("Hand tools and Power tools", "Hand Tools Power driven"),
    ("Housekeeping", "Housekeeping"),
    ("Welfare facilities and Labour accommodation", "Labour Welfare at work place"),
    ("Welfare facilities and Labour accommodation", "Labour Hutment"),
    ("Lifting Appliances and Gear", "Lifting Appliances"),
    ("Lifting Appliances and Gear", "Lifting Tools and Tackles"),
    ("Earth Moving Equipment", "Earth moving equipment"),
    ("Piling Work", "Piling Work"),
    ("Excavation & Confined space job", "Excavation"),
    ("Fire Prevention / Protection", "Firefighting"),
    ("General safety", "Working near or on waterbody"),
    ("Welding cutting grinding", "Welding"),
    ("Welding cutting grinding", "Gas cutting"),
    ("Welding cutting grinding", "Grinding"),
    ("Vehicles", "Vehicle"),
    ("Excavation & Confined space job", "Confined space entry"),
    ("General safety", "Manual material handling"),
    ("Environment Aspects", "Environment abuse"),
    ("General safety", "Working condition"),
    ("General safety", "Risky Behaviour"),
    ("General safety", "Praiseworthy observation"),
    ("General safety", "Other issue"),
]

SPECIFIC_CONCERNS_MAP = {
    "Personal Protective equipment": [
        "Damaged / Worn out PPE", "Safety Helmet not worn", "Safety shoes not worn",
        "Full body harness not worn", "Improper anchoring of safety harness",
        "Safety jacket not worn", "No PPEs worn at all", "Job specific PPEs not worn",
    ],
    "Floor edge (Work at height)": [
        "Not barricaded / inadequately barricaded",
        "Safety net to protect fall of person is not provided",
        "Safety net to protect fall of material is not provided",
        "Damaged barricade. Needs repair", "No warning signage",
        "Safety net is either damaged or improperly fixed.",
        "Barricaded using unapproved means", "Debris accumulation on safety nets",
    ],
    "Floor opening (Work at height)": [
        "Not barricaded / inadequately barricaded", "Barricaded using unapproved means",
        "Safety net is required but not fixed", "Uncovered pits", "No warning signage",
        "Safety net is either damaged or improperly fixed.",
        "Damaged barricade. Needs repair", "Debris accumulation on safety nets",
    ],
    "Machinery": [
        "Unguarded rotating part", "Damaged", "Accessories missing",
        "Safety gadgets missing", "Faulty", "No certification obtained",
        "Manipulated / Improvised / tampered",
    ],
    "Scaffold (Work at height)": [
        "Substandard materials used for construction", "Members missing", "Not in plumb",
        "Bad ground condition", "Improper footing", "Inadequately supported",
        "Appropriate tag missing", "Incomplete but still in use", "Close to power line",
    ],
    "Ladder (Work at height)": [
        "Unsupported", "Damaged", "Bad ground condition",
        "Manipulated / Improvised / tampered", "Unmanned",
        "Inappropriate for the task", "Wrongly positioned",
    ],
    "Work platform (Work at height)": [
        "Substandard materials used for construction", "Members missing", "Handrails missing",
        "Inadequately secured", "No lifeline provided for anchorage",
        "Lifeline is not as per standard", "Inadequately supported", "Inadequately decked",
        "No approach path", "Incomplete", "Close to power line", "Appropriate tag missing",
    ],
    "Electrical": [
        "No plug top / Bare wires inserted", "Substandard plug top",
        "RCCB not available / substandard", "RCCB inadequate in number",
        "Substandard distribution box / extension board", "Damaged power cable",
        "Cable in contact with sharp edges", "Substandard cable Joint",
        "Poor illumination", "No fire fighting device", "Improper cable routing",
        "No rescue kit", "No earthing", "No rubber mat",
        "Use of unapproved wires", "Substandard / inadequate insulation",
    ],
    "Hand Tools Manual": [
        "Wrong tool used for the job", "Insulation worn out", "Damaged / Worn out",
        "Faulty", "Manipulated / Improvised / tampered",
    ],
    "Hand Tools Power driven": [
        "Unguarded rotating part", "Electrical protection deteriorated",
        "Accessories missing /mismatch", "Damaged / Worn out",
        "Faulty", "Manipulated / Improvised / tampered",
    ],
    "Housekeeping": [
        "Improper stacking of materials", "Unclean work place",
        "Chemicals / Fuels not stored as per recommendations", "Materials found scattered",
        "Protruding rods or nails", "Sharp objects in exposed condition",
        "Obstacles / Obstruction", "Danger of slipping", "Danger of tripping",
        "Storage of inflammables haphazardly", "Walkways / Aisles are cluttered",
        "Walkways / Aisles are not designated or marked",
        "Severe undulation on the road/path", "Loose materials at floor edge",
        "Debris accumulation in non-designated place",
        "Insufficient numbers of safety posters and sign boards", "Stagnant water",
    ],
    "Labour Welfare at work place": [
        "Insufficient/No provision of drinking water",
        "Insufficient/No provision of shelter at work place",
        "Insufficient/No urinals at work place", "Insufficient number of First Aid box",
        "Stretcher unavailable / insufficient", "Resting in non-designated area",
        "Inappropriate attire at work",
    ],
    "Labour Hutment": [
        "Hole in walls and roof", "Substandard flooring", "Poor illumination",
        "Poor housekeeping inside rooms", "Constructed using inflammable material",
        "Substandard bedding", "Inadequate ventilation",
        "Sharp objects in exposed condition", "Unapproved materials used for construction",
        "Unapproved design", "Insufficient number of First Aid box",
        "Stretcher not available", "Unapproved electrical distribution",
        "Substandard sanitation system", "Storage of inflammables haphazardly",
        "Lack of hygiene", "Stagnant water", "Substandard Cooking facilities",
        "Improper stacking of materials",
    ],
    "Lifting Appliances": [
        "Unguarded rotating part", "Damaged", "Accessories missing",
        "Safety gadgets missing", "Faulty", "No certification obtained", "Manipulated",
    ],
    "Lifting Tools and Tackles": [
        "Unguarded", "Damaged", "Faulty", "No certification obtained", "Manipulated",
    ],
    "Earth moving equipment": [
        "Missing accessories", "Damaged", "Faulty", "No certification obtained", "Manipulated",
    ],
    "Piling Work": [
        "Damaged lifting tools and tackles", "Abandoned pile pits",
        "Work place unapproachable", "Faulty Rig", "Faulty winch",
        "Missing accessories", "Unguarded rotating part",
        "No certification obtained", "Manipulated / Improvised / tampered",
    ],
    "Excavation": [
        "No hard barricade (>1.2 m)", "No soft barricade (<1.2 m)",
        "Improper access", "Undercut", "No sloping or benching",
        "No shoring / struting", "No caution boards",
        "Materials close to edge of excavated area",
        "Inadequate illumination", "Water accumulation",
    ],
    "Firefighting": [
        "Inadequate number of extinguisher", "Inappropriate extinguisher",
        "Damaged fire fighting equipment", "Incorrect positioning of extinguishers",
    ],
    "Working near or on waterbody": [
        "Persons not physically fit", "Machinery manipulated / Improvised / tampered",
        "No rescue equipment", "No supervision", "No barricade", "Faulty pontoon / vessel",
    ],
    "Welding": [
        "Poor condition of machine", "Poor condition of cable", "Safety gadgets missing",
        "Job specific PPEs not worn", "Improper accessories", "No lugs", "No earthing",
    ],
    "Gas cutting": [
        "No Flash back arrestor", "Poor hose condition", "Mismatch of hose colour",
        "Job specific PPEs not worn", "Poor gauge condition",
        "Improper positioning of gas cylinders", "Hotwork near combustible materials",
    ],
    "Grinding": [
        "Damaged cutting wheel", "Job specific PPEs not worn",
        "Inappropriate cutting wheel", "Unguarded rotating part", "Manipulated machine",
    ],
    "Vehicle": [
        "Missing accessories", "Damaged", "Faulty", "No certification obtained", "Manipulated",
    ],
    "Confined space entry": [
        "Entry without permission", "Persons not physically fit", "No supervision",
        "Entry without information", "Safety gadgets missing / inadequate",
        "Entry without testing", "Entry without precaution",
    ],
    "Manual material handling": [
        "Wrong process of lifting load", "Insufficient persons", "Cluttered route",
        "Head loading", "Wrong posture of persons", "Excess load being lifted",
    ],
    "Environment abuse": [
        "Possibility of Oil spillage", "Possibility of Chemical spillage",
        "Excessive dusty conditions", "Excessively noisy", "Muddy conditions of roads",
    ],
    "Working condition": [
        "Excessively Noisy", "Excessively dusty", "Undulated floor condition",
        "Poorly Illuminated / Dark Work place", "Presence of chemical or fumes",
        "Damp and Moist", "Hot and humid", "Congested",
    ],
    "Risky Behaviour": [
        "Driving/ Operating : recklessly/wrongly", "Driving/ Operating without authority",
        "Driving / Operating without permission", "Horse playing / Teasing",
        "Using unsafe alternative device / tools", "Standing / moving inside swing area",
        "Standing / moving under the load", "Adopting unsafe technique to perform task",
        "Adopting incorrect lifting techniques", "Violating safety instructions",
        "Using wrong tools for the task", "Smoking",
        "Working under influence of alcohol", "Abusing Personal Protective Equipment",
        "Person either below 18 or above 58",
    ],
    "Praiseworthy observation": [],
    "Other issue": [],
}

POSSIBLE_OUTCOMES = [
    "Person may fall on the same level", "Person may fall from height",
    "Person may get hit", "Person may drown",
    "Person may receive electrical shock", "Person may get burns",
    "Person may lose eyesight", "Person may choke",
    "Person may get runover by vehicle", "Person may fall ill",
    "Person may not be visible from distance", "Person may hurt his/her feet",
    "Person may get entangled", "Materials may roll down",
    "Materials may fall from height", "Danger of explosion",
    "Soil pollution", "Water pollution", "Air pollution",
    "Project may look shabby", "Fire may occur",
    "Chances of property damage", "Soil subsidence",
    "Violation of legal obligation", "Emergency evacuation may be difficult",
    "Vehicular collision / topple / subsidence", "No chance of anything happening",
]

TARGET_DATES = [
    "Stop the job. Rectify immediately",
    "Rectify before the job starts",
    "Rectify within 2 days",
    "Rectify within 4 days",
    "Rectify within 7 days",
    "Continue the good practice",
]

VIOLATIONS = [
    "Job procedure unavailable", "Job procedure inaccurate", "Job procedure violated",
    "Job procedure not explained properly", "Instructions not understood",
    "Supervision not present", "Supervision not competent", "Supervision not strong",
    "Workman not competent", "Workman physically unfit", "Lack of training on the task",
    "No maintenance / servicing", "Irregular maintenance / servicing",
    "Disobeying the supervision", "Noticed but moved on",
    "Clause not mentioned in contract", "Contract clause not enforced",
    "Confusion / Misunderstanding", "Poor perception of danger",
    "Overconfidence", "Haste and Hurry", "Unergonomic workplace", "Unhygienic work place",
]

ROOT_CAUSE_CATEGORIES = [
    "Procedures", "Conflicting targets", "Communication", "Hardware", "Design",
    "Environmental Conditions", "Training", "Organization", "Maintenance", "Discipline",
]

ROOT_CAUSE_SPECIFICS = {
    "Procedures": [
        "Inadequate team that developed the procedure.",
        "Inadequate feedback about the procedures.", "Bypassing procedure",
        "Lack of standardization.",
        "Personnel not informed about procedures.", "Lack of Procedure",
        "Lack of requirement to have task procedure.",
    ],
    "Conflicting targets": [
        "Conflict between production and safe working practices.",
        "Conflict between financial priorities and safe working practices.",
        "Conflict between social and domestic priorities on safe working practices.",
        "Conflict between individual priorities and safe working practices.",
    ],
    "Communication": [
        "Language problem / Cultural barriers.", "Lack of clear lines of communications.",
        "Overruling the supervision", "Inadequate Feedback.",
        "No standardization of information formats.",
        "Inadequate communication between design and user, allowing substandard conditions to remain.",
        "Lines of communication overloaded.",
    ],
    "Hardware": [
        "Inadequate specification of equipment.", "Wrong components purchased / used.",
        "Equipment may not be available leading to improvisation.",
        "Defective equipments.", "Age of equipment compared to life expectancy.",
        "Servicing / Maintenance are not carried out in time.",
    ],
    "Design": [
        "Design not fully understood to operate safely.",
        "Poor specifications leading to substandard material and inherent weakness in the system.",
        "Specifications not complied with requirement.", "Not ergonomically designed.",
    ],
    "Environmental Conditions": [
        "Poor morals caused by a number of situations e.g. unfair enforcement of rules, weak discipline etc.",
        "Physical deterioration, caused by long working hours, subject to undue pressure.",
        "Staff unable to cope up or respond to emergency or unusual situation.",
        "Information not correct or supplied.",
        "Poor co-ordination between departments causing interface problems.",
        "Circumstantial factors such as domestic pressures, homesickness etc.",
    ],
    "Training": [
        "Ineffective pre-employment selection.",
        "Poor education not compatible with job requirements.",
        "No structured planning of training programmes.",
        "No assessment of training effectiveness.",
        "Ineffective training - Lack of subject Knowledge",
        "Training not appropriate for the personnel selected.",
        "Lack of Anticipation or foresightedness",
    ],
    "Organization": [
        "Inadequately defined departments or parts of organization.",
        "Unclear accountability, responsibility or delegation structure.",
        "Inadequate definitions of objectives and co-ordination of project and tasks.",
        "Frequent reorganization of departments.", "Too much bureaucracy and rigidity.",
    ],
    "Maintenance": [
        "Inadequate maintenance programme.", "Shortage of maintenance personnel.",
        "Financial or time constraints.",
    ],
    "Discipline": [
        "Ergonomic consideration, human limitation e.g., size and strength of individuals.",
        "Looking the other way",
        "Financial and other constraints during design/execution phase.",
        "Poor discipline and enforcement",
    ],
}

# ── Style helpers ─────────────────────────────────────────────────────────────

def hdr_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

THIN = Side(style="thin", color="AAAAAA")
MED  = Side(style="medium", color="555555")

def thin_border():
    return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def med_border():
    return Border(left=MED, right=MED, top=MED, bottom=MED)

HDR_FONT   = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
BODY_FONT  = Font(size=10, name="Calibri")
NOTE_FONT  = Font(size=9, italic=True, color="666666", name="Calibri")
TITLE_FONT = Font(bold=True, size=14, name="Calibri", color="1F4E79")

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

# Column header → (fill hex, width)
COLUMNS = [
    # ── Identification ────────────────────────────────────────────────────────
    ("Observation ID",          "1F4E79", 20),   # A  – e.g. OBS-2024-001
    ("Observation Date\n(DD-MM-YYYY)", "1F4E79", 18),  # B
    ("Observation Time\n(HH:MM)", "1F4E79", 16),  # C
    # ── Location ─────────────────────────────────────────────────────────────
    ("Project Name",            "2E75B6", 25),   # D
    ("Building / Block",        "2E75B6", 20),   # E
    ("Floor / Level",           "2E75B6", 18),   # F
    ("Exact Location",          "2E75B6", 25),   # G
    # ── People ───────────────────────────────────────────────────────────────
    ("Observer Name",           "375623", 22),   # H
    ("Contractor Name",         "375623", 22),   # I
    ("To Be Rectified By",      "375623", 22),   # J
    # ── Observation Classification ────────────────────────────────────────────
    ("Category",                "7030A0", 30),   # K  – dropdown
    ("Core Concern",            "7030A0", 35),   # L  – dropdown
    ("Specific Concern",        "7030A0", 35),   # M  – dropdown
    ("Specific Concern (Custom Text)", "7030A0", 35),  # N  – free text override
    # ── Risk ─────────────────────────────────────────────────────────────────
    ("Possible Outcome",        "C55A11", 40),   # O  – dropdown
    ("Severity\n(1=Low … 5=High)", "C55A11", 18),   # P  – 1-5
    ("Probability\n(1=Rare … 5=Almost Certain)", "C55A11", 18),  # Q  – 1-5
    ("Risk Factor\n(Severity × Probability)", "C55A11", 18),  # R  – formula
    ("Risk Level\n(Low/Medium/High)", "C55A11", 18),  # S  – formula
    # ── Root Cause ────────────────────────────────────────────────────────────
    ("Root Cause Category",     "843C0C", 30),   # T  – dropdown
    ("Root Cause Specific",     "843C0C", 45),   # U  – dropdown
    # ── Action ────────────────────────────────────────────────────────────────
    ("Violation / Non-Conformance", "1F4E79", 45),  # V  – dropdown
    ("Target Date Type",        "1F4E79", 35),   # W  – dropdown
    ("Target Date Actual\n(DD-MM-YYYY)", "1F4E79", 20),  # X
    ("Status\n(Open/Closed)",   "1F4E79", 15),   # Y  – dropdown
]

# ── Build workbook ─────────────────────────────────────────────────────────────

wb = openpyxl.Workbook()

# ── 1. INSTRUCTIONS sheet ────────────────────────────────────────────────────
ins = wb.active
ins.title = "Instructions"
ins.sheet_view.showGridLines = False

ins.merge_cells("A1:D1")
ins["A1"] = "Safety Observation System — Data Migration Template"
ins["A1"].font = TITLE_FONT
ins["A1"].alignment = CENTER
ins.row_dimensions[1].height = 32

instructions = [
    ("", ""),
    ("PURPOSE", "Use this template to export historical safety observation records for import into the Safety Observation App."),
    ("", ""),
    ("HOW TO USE", "1. Fill in the 'Observations' sheet — one row per observation."),
    ("", "2. Use the dropdown lists in each column where provided. Refer to the 'Reference' sheet for all valid values."),
    ("", "3. Dates must be in DD-MM-YYYY format (e.g., 25-01-2024)."),
    ("", "4. Times must be in HH:MM 24-hour format (e.g., 14:30)."),
    ("", "5. Severity and Probability values are integers 1–5 only."),
    ("", "6. Risk Factor and Risk Level are auto-calculated — do NOT edit those columns."),
    ("", "7. Leave a cell blank if the information is not available — do not enter 'N/A' or dashes."),
    ("", "8. Images will be uploaded separately by the admin directly in the app after import."),
    ("", ""),
    ("MANDATORY FIELDS", "• Observation Date\n• Project Name\n• Observer Name\n• Category\n• Core Concern\n• Severity\n• Probability\n• Status"),
    ("", ""),
    ("RISK MATRIX", "Risk Factor = Severity × Probability\n• Low  : ≤ 4\n• Medium: 5 – 12\n• High : ≥ 15"),
    ("", ""),
    ("NOTES", "• Observation ID will be auto-generated by the system if left blank.\n• If Specific Concern is not in the dropdown, enter it in the 'Specific Concern (Custom Text)' column.\n• 'To Be Rectified By' is a free-text name (contractor / person responsible).\n• Contractor Name must match an existing contractor in the system."),
]

for r, (col1, col2) in enumerate(instructions, start=2):
    ins.cell(row=r, column=1).value = col1
    ins.cell(row=r, column=1).font = Font(bold=True, size=10, name="Calibri", color="1F4E79")
    ins.cell(row=r, column=1).alignment = Alignment(vertical="top", wrap_text=True)
    ins.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    ins.cell(row=r, column=2).value = col2
    ins.cell(row=r, column=2).font = BODY_FONT
    ins.cell(row=r, column=2).alignment = Alignment(vertical="top", wrap_text=True)
    ins.row_dimensions[r].height = max(15, col2.count("\n") * 15 + 15)

ins.column_dimensions["A"].width = 22
ins.column_dimensions["B"].width = 80
ins.column_dimensions["C"].width = 5
ins.column_dimensions["D"].width = 5

# ── 2. REFERENCE sheet ────────────────────────────────────────────────────────
ref = wb.create_sheet("Reference")
ref.sheet_view.showGridLines = False
ref.sheet_state = "visible"

ref_sections = [
    ("Categories",              CATEGORIES),
    ("Core Concerns",           [cc for _, cc in CORE_CONCERNS]),
    ("Possible Outcomes",       POSSIBLE_OUTCOMES),
    ("Target Date Types",       TARGET_DATES),
    ("Violations",              VIOLATIONS),
    ("Root Cause Categories",   ROOT_CAUSE_CATEGORIES),
    ("Status Values",           ["Open", "Closed"]),
    ("Severity Labels",         ["1 - Very Low", "2 - Low", "3 - Medium", "4 - High", "5 - Very High"]),
    ("Probability Labels",      ["1 - Rare", "2 - Unlikely", "3 - Possible", "4 - Likely", "5 - Almost Certain"]),
]

# Write each section side-by-side in columns
col_offset = 1
for section_title, values in ref_sections:
    # Header
    hdr_cell = ref.cell(row=1, column=col_offset)
    hdr_cell.value = section_title
    hdr_cell.font = Font(bold=True, size=10, name="Calibri", color="FFFFFF")
    hdr_cell.fill = hdr_fill("1F4E79")
    hdr_cell.alignment = CENTER
    hdr_cell.border = thin_border()

    for i, val in enumerate(values, start=2):
        c = ref.cell(row=i, column=col_offset)
        c.value = val
        c.font = BODY_FONT
        c.alignment = LEFT
        c.border = thin_border()

    ref.column_dimensions[get_column_letter(col_offset)].width = 40
    col_offset += 2  # gap column between sections

# Specific concerns — write each core concern block as a sub-table
sc_col = col_offset
sc_hdr = ref.cell(row=1, column=sc_col)
sc_hdr.value = "Specific Concerns by Core Concern"
sc_hdr.font = Font(bold=True, size=10, name="Calibri", color="FFFFFF")
sc_hdr.fill = hdr_fill("7030A0")
sc_hdr.alignment = CENTER
sc_hdr.border = thin_border()
ref.merge_cells(start_row=1, start_column=sc_col, end_row=1, end_column=sc_col + 1)

sc_row = 2
for cc_name, specifics in SPECIFIC_CONCERNS_MAP.items():
    # Core concern title
    cc_cell = ref.cell(row=sc_row, column=sc_col)
    cc_cell.value = cc_name
    cc_cell.font = Font(bold=True, size=9, name="Calibri", color="FFFFFF")
    cc_cell.fill = hdr_fill("7030A0")
    cc_cell.alignment = LEFT
    cc_cell.border = thin_border()
    ref.merge_cells(start_row=sc_row, start_column=sc_col, end_row=sc_row, end_column=sc_col + 1)
    sc_row += 1
    for sp in specifics:
        sp_cell = ref.cell(row=sc_row, column=sc_col)
        sp_cell.value = sp
        sp_cell.font = BODY_FONT
        sp_cell.alignment = LEFT
        sp_cell.border = thin_border()
        ref.merge_cells(start_row=sc_row, start_column=sc_col, end_row=sc_row, end_column=sc_col + 1)
        sc_row += 1
    sc_row += 1  # blank row between groups

ref.column_dimensions[get_column_letter(sc_col)].width = 50

# Root cause specifics — next block
rcs_col = sc_col + 3
rcs_hdr = ref.cell(row=1, column=rcs_col)
rcs_hdr.value = "Root Cause Specifics by Category"
rcs_hdr.font = Font(bold=True, size=10, name="Calibri", color="FFFFFF")
rcs_hdr.fill = hdr_fill("843C0C")
rcs_hdr.alignment = CENTER
rcs_hdr.border = thin_border()
ref.merge_cells(start_row=1, start_column=rcs_col, end_row=1, end_column=rcs_col + 1)

rcs_row = 2
for cat_name, specifics in ROOT_CAUSE_SPECIFICS.items():
    cat_cell = ref.cell(row=rcs_row, column=rcs_col)
    cat_cell.value = cat_name
    cat_cell.font = Font(bold=True, size=9, name="Calibri", color="FFFFFF")
    cat_cell.fill = hdr_fill("843C0C")
    cat_cell.alignment = LEFT
    cat_cell.border = thin_border()
    ref.merge_cells(start_row=rcs_row, start_column=rcs_col, end_row=rcs_row, end_column=rcs_col + 1)
    rcs_row += 1
    for sp in specifics:
        sp_cell = ref.cell(row=rcs_row, column=rcs_col)
        sp_cell.value = sp
        sp_cell.font = BODY_FONT
        sp_cell.alignment = LEFT
        sp_cell.border = thin_border()
        ref.merge_cells(start_row=rcs_row, start_column=rcs_col, end_row=rcs_row, end_column=rcs_col + 1)
        rcs_row += 1
    rcs_row += 1

ref.column_dimensions[get_column_letter(rcs_col)].width = 70

# ── 3. OBSERVATIONS sheet ─────────────────────────────────────────────────────
obs = wb.create_sheet("Observations")
obs.sheet_view.showGridLines = False
obs.freeze_panes = "A3"     # freeze title + header rows

# Row 1 — banner
obs.merge_cells("A1:Y1")
obs["A1"] = "Safety Observation System — Data Migration"
obs["A1"].font = Font(bold=True, size=13, name="Calibri", color="FFFFFF")
obs["A1"].fill = hdr_fill("1F4E79")
obs["A1"].alignment = CENTER
obs.row_dimensions[1].height = 28

# Row 2 — section labels
SECTION_SPANS = [
    ("Identification",              "A2:C2", "1F4E79"),
    ("Location",                    "D2:G2", "2E75B6"),
    ("People",                      "H2:J2", "375623"),
    ("Observation Classification",  "K2:N2", "7030A0"),
    ("Risk Assessment",             "O2:S2", "C55A11"),
    ("Root Cause",                  "T2:U2", "843C0C"),
    ("Action",                      "V2:Y2", "1F4E79"),
]
for label, span, color in SECTION_SPANS:
    obs.merge_cells(span)
    start_cell = span.split(":")[0]
    cell = obs[start_cell]
    cell.value = label
    cell.font = Font(bold=True, size=10, name="Calibri", color="FFFFFF")
    cell.fill = hdr_fill(color)
    cell.alignment = CENTER
    cell.border = med_border()

obs.row_dimensions[2].height = 20

# Row 3 — column headers
for col_idx, (header, color, width) in enumerate(COLUMNS, start=1):
    cell = obs.cell(row=3, column=col_idx)
    cell.value = header
    cell.font = Font(bold=True, size=9, name="Calibri", color="FFFFFF")
    cell.fill = hdr_fill(color)
    cell.alignment = CENTER
    cell.border = thin_border()
    obs.column_dimensions[get_column_letter(col_idx)].width = width

obs.row_dimensions[3].height = 36

# ── Sample rows ───────────────────────────────────────────────────────────────
SAMPLE_ROWS = [
    [
        "OBS-2024-001", "15-01-2024", "09:30",
        "Site A - Tower Block 1", "Tower A", "5th Floor", "North staircase landing",
        "Rajesh Kumar", "ABC Contractors Pvt Ltd", "Site Supervisor",
        "Work at Height", "Scaffold (Work at height)", "Members missing", "",
        "Person may fall from height",
        4, 3, None, None,
        "Procedures", "Lack of Procedure",
        "Job procedure violated",
        "Rectify within 2 days", "17-01-2024",
        "Closed",
    ],
    [
        "OBS-2024-002", "16-01-2024", "14:00",
        "Site A - Tower Block 1", "Tower B", "3rd Floor", "Electrical room entrance",
        "Priya Sharma", "XYZ Electricals", "Electrical Supervisor",
        "Electrical", "Electrical", "No earthing", "",
        "Person may receive electrical shock",
        5, 4, None, None,
        "Maintenance", "Inadequate maintenance programme.",
        "No maintenance / servicing",
        "Stop the job. Rectify immediately", "16-01-2024",
        "Open",
    ],
    [
        "", "20-01-2024", "11:15",
        "Site B - Residential Block", "Block C", "Ground Floor", "Material storage yard",
        "Mohan Das", "PQR Construction", "",
        "Housekeeping", "Housekeeping", "Improper stacking of materials", "",
        "Materials may fall from height",
        3, 2, None, None,
        "Organization", "Too much bureaucracy and rigidity.",
        "Supervision not present",
        "Rectify within 4 days", "24-01-2024",
        "Open",
    ],
]

ALT_FILL = PatternFill("solid", fgColor="F2F7FF")

for r_idx, row_data in enumerate(SAMPLE_ROWS, start=4):
    fill = ALT_FILL if r_idx % 2 == 0 else None
    for c_idx, value in enumerate(row_data, start=1):
        cell = obs.cell(row=r_idx, column=c_idx)
        # Skip risk factor (col 18) and risk level (col 19) — formula rows
        if c_idx in (18, 19):
            continue
        cell.value = value
        cell.font = BODY_FONT
        cell.alignment = LEFT
        cell.border = thin_border()
        if fill:
            cell.fill = fill

# ── Formula columns (Risk Factor = P*Q, Risk Level) for sample rows + template rows ──
DATA_START = 4
DATA_END   = 1003   # support 1000 data rows

for r in range(DATA_START, DATA_END + 1):
    # Risk Factor = Severity × Probability  (cols P=16, Q=17 → R=18)
    rf_cell = obs.cell(row=r, column=18)
    rf_cell.value = f"=IF(AND(ISNUMBER(P{r}),ISNUMBER(Q{r})),P{r}*Q{r},\"\")"
    rf_cell.font = Font(size=10, name="Calibri", italic=True, color="555555")
    rf_cell.alignment = CENTER
    rf_cell.border = thin_border()

    # Risk Level
    rl_cell = obs.cell(row=r, column=19)
    rl_cell.value = (
        f'=IF(R{r}="","",IF(R{r}<=4,"Low",IF(R{r}<=12,"Medium","High")))'
    )
    rl_cell.font = Font(size=10, name="Calibri", italic=True, color="555555")
    rl_cell.alignment = CENTER
    rl_cell.border = thin_border()

# ── Data validations ─────────────────────────────────────────────────────────
def make_list_dv(formula, rows_range):
    dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
    dv.sqref = rows_range
    dv.showErrorMessage = True
    dv.errorTitle = "Invalid Value"
    dv.error = "Please select a value from the dropdown list."
    return dv

# Write lookup lists to a hidden sheet so validation formulas work
lkp = wb.create_sheet("_Lookups")
lkp.sheet_state = "hidden"

def write_lookup_col(ws, col, values, header=None):
    start = 1
    if header:
        ws.cell(row=1, column=col).value = header
        start = 2
    for i, v in enumerate(values, start=start):
        ws.cell(row=i, column=col).value = v
    return start, start + len(values) - 1

# Col A: categories
_, cat_end = write_lookup_col(lkp, 1, CATEGORIES)
# Col B: core concerns
_, cc_end = write_lookup_col(lkp, 2, [cc for _, cc in CORE_CONCERNS])
# Col C: possible outcomes
_, po_end = write_lookup_col(lkp, 3, POSSIBLE_OUTCOMES)
# Col D: target dates
_, td_end = write_lookup_col(lkp, 4, TARGET_DATES)
# Col E: violations
_, vio_end = write_lookup_col(lkp, 5, VIOLATIONS)
# Col F: root cause categories
_, rcc_end = write_lookup_col(lkp, 6, ROOT_CAUSE_CATEGORIES)
# Col G: status
_, st_end = write_lookup_col(lkp, 7, ["Open", "Closed"])

# Add validations to Observations sheet
obs.add_data_validation(make_list_dv(f"_Lookups!$A$1:$A${cat_end}",  f"K4:K{DATA_END}"))
obs.add_data_validation(make_list_dv(f"_Lookups!$B$1:$B${cc_end}",   f"L4:L{DATA_END}"))
obs.add_data_validation(make_list_dv(f"_Lookups!$C$1:$C${po_end}",   f"O4:O{DATA_END}"))
obs.add_data_validation(make_list_dv(f"_Lookups!$D$1:$D${td_end}",   f"W4:W{DATA_END}"))
obs.add_data_validation(make_list_dv(f"_Lookups!$E$1:$E${vio_end}",  f"V4:V{DATA_END}"))
obs.add_data_validation(make_list_dv(f"_Lookups!$F$1:$F${rcc_end}",  f"T4:T{DATA_END}"))
obs.add_data_validation(make_list_dv(f"_Lookups!$G$1:$G${st_end}",   f"Y4:Y{DATA_END}"))

# Severity / Probability : 1-5
sev_dv = DataValidation(type="whole", operator="between", formula1="1", formula2="5",
                         allow_blank=True)
sev_dv.sqref = f"P4:Q{DATA_END}"
sev_dv.showErrorMessage = True
sev_dv.errorTitle = "Invalid Value"
sev_dv.error = "Enter a whole number between 1 and 5."
obs.add_data_validation(sev_dv)

# ── Blank template rows (rows 4-10 pre-formatted) ────────────────────────────
for r in range(DATA_START + len(SAMPLE_ROWS), DATA_START + len(SAMPLE_ROWS) + 7):
    fill = ALT_FILL if r % 2 == 0 else None
    for c in range(1, len(COLUMNS) + 1):
        if c in (18, 19):
            continue
        cell = obs.cell(row=r, column=c)
        cell.font = BODY_FONT
        cell.alignment = LEFT
        cell.border = thin_border()
        if fill:
            cell.fill = fill

# ── Tab order ─────────────────────────────────────────────────────────────────
wb.active = obs   # open on Observations sheet

# ── Save ──────────────────────────────────────────────────────────────────────
output_path = "Safety_Observation_Migration_Template.xlsx"
wb.save(output_path)
print(f"Template saved: {output_path}")
